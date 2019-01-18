from datetime import datetime, timedelta, date
from itertools import groupby
from typing import List

from flask_injector import inject
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.styles import Font, Border, Side, Alignment, colors

from nisse.services.user_service import UserService
from nisse.services.vacation_service import VacationService
from nisse.utils.date_helper import *
from nisse.utils.string_helper import *


class XlsxDocumentService(object):

    @inject
    def __init__(self, user_service: UserService, vacation_service: VacationService):
        self.vacation_service = vacation_service
        self.user_service = user_service

    font_red = Font(color=colors.RED)
    font_orange = Font(color='00FF9900')
    font_red_bold = Font(color=colors.RED, bold=True)
    font_bold = Font(bold=True)
    top_border = Border(top=Side(style='thin'))
    alignment_right = Alignment(horizontal='right')
    alignment_top = Alignment(vertical='top')

    def save_report(self, file_path, date_from, date_to, time_entries):
        """
        Creates report and saves it into xlsx file
        :param file_path: file destination
        :param date_from: start date for report
        :param date_to: end date for report
        :param time_entries: collection time entries
        :param vacation_data: list of vacations
        :return:
        """

        wb = Workbook()

        time_entries = sorted(time_entries, key=lambda te: get_user_name(te.user))
        by_user = groupby(time_entries, key=lambda te: te.user_id)

        first_sheet = True
        for user_id, group in by_user:

            user: User = self.user_service.get_user_by_id(user_id)
            if not user:
                continue

            group_sorted = list(group)
            vacation_data = self.vacation_service.get_vacations_by_dates(user.user_id, date_from, date_to)

            first_name = get_user_name(user)
            sheet = None
            if first_sheet:
                sheet = wb.active
                sheet.title = first_name
                first_sheet = False
            else:
                sheet = wb.create_sheet(first_name)

            sheet.column_dimensions['A'].width = 12
            sheet.column_dimensions['B'].width = 10
            sheet.column_dimensions['C'].width = 20
            sheet.column_dimensions['D'].width = 60

            self.put_text(sheet['A1'], "Date", font=self.font_bold)
            self.put_text(sheet['B1'], "Duration", font=self.font_bold)
            self.put_text(sheet['C1'], "Project", font=self.font_bold)
            self.put_text(sheet['D1'], "Comment", font=self.font_bold)

            i = 1
            i_start = i + 1
            total_basic = 0
            total_deficit = 0
            total_overtime = 0
            for day in date_range(datetime.strptime(date_from, "%Y-%m-%d").date(),
                                   datetime.strptime(date_to, "%Y-%m-%d").date() + timedelta(days=1)):

                tes = list(filter(lambda te: te.report_date == day, group_sorted))
                time_reported = sum(te.duration for te in tes)

                self.put_text(sheet['A' + str(i + 1)], format_date(day),
                              font=self.get_date_color(day, vacation_data), alignment=self.alignment_top)

                for te in tes:
                    i += 1
                    self.put_time(sheet['B' + str(i)], te.duration)
                    self.put_time(sheet['C' + str(i)], te.project.name)
                    self.put_text(sheet['D' + str(i)], te.comment)

                if not len(tes):
                    i += 1
                    self.put_time(sheet['B' + str(i)], 0)

                to_merge = len(tes) - 1 if len(tes) else 0
                sheet.merge_cells("A" + str(i - to_merge) + ":A" + str(i))

                basic = 0
                deficit = 0
                if not is_weekend(day) and not XlsxDocumentService.is_vacation_day(day, vacation_data):
                    if not time_reported:
                        time_reported = 0
                    basic = time_reported if time_reported <= 8 else 8
                    deficit = 8 - basic if time_reported <= 8 else 0
                overtime = time_reported - basic
                total_overtime += overtime
                total_deficit += deficit
                total_basic += basic

            total_overtime_deficit = total_overtime - total_deficit if (total_overtime - total_deficit) > 0 else 0
            total_basic_deficit = total_basic + total_deficit if total_overtime_deficit > 0 else total_basic + total_overtime
            total_deficit = total_deficit - total_overtime if (total_deficit - total_overtime) > 0 else 0

            i += 1
            XlsxDocumentService.put_time(sheet['B' + str(i)], str("=SUM(B" + str(i_start) + ":B" + str(i - 1) + ")"),
                          font=self.font_bold)
            XlsxDocumentService.put_text(sheet['A' + str(i + 2)], "Overtime:", font=self.font_bold, alignment=self.alignment_right)
            XlsxDocumentService.put_time(sheet['B' + str(i + 2)], total_overtime_deficit, self.font_bold)
            XlsxDocumentService.put_text(sheet['A' + str(i + 3)], "Basic hours:", font=self.font_bold, alignment=self.alignment_right)
            XlsxDocumentService.put_time(sheet['B' + str(i + 3)], total_basic_deficit, self.font_bold)
            XlsxDocumentService.put_text(sheet['A' + str(i + 4)], "Deficit:", font=self.font_bold,
                                         alignment=self.alignment_right)
            XlsxDocumentService.put_time(sheet['B' + str(i + 4)], total_deficit, self.font_bold)

        wb.save(file_path)

    def get_date_color(self, day: date, vacation_list):
        if is_weekend(day):
            return self.font_red
        elif XlsxDocumentService.is_vacation_day(day, vacation_list):
            return self.font_orange
        else:
            return None

    @staticmethod
    def is_vacation_day(day: date, vacation_list: List[Vacation]):
        for vacation in vacation_list:
            if vacation.start_date <= day <= vacation.end_date:
                return True
        return False

    @staticmethod
    def put_time(cell: Cell, duration, font=None, border=None, alignment=None):
        cell = XlsxDocumentService.put_text(cell, duration, font, border, alignment)
        cell.number_format = '0.00'
        return cell

    @staticmethod
    def put_text(cell: Cell, text, font=None, border=None, alignment=None):
        cell.value = text
        if font:
            cell.font = font
        if border:
            cell.border = border
        if alignment:
            cell.alignment = alignment
        return cell
